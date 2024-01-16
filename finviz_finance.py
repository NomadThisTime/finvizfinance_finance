### import packages

from finvizfinance.quote import finvizfinance
from openpyxl import load_workbook

### loading workbook

wb = load_workbook('finvizfinance_workbook.xlsx')
ws = wb.active

### defining maximum column and row

def get_maximum_cols():
    for i in range(2, 20000):
        if ws.cell(row=1, column= i).value == None:
            max_col = i
            break
    return max_col-1

def get_maximum_rows():
    for i in range(2, 20000):
        if ws.cell(row=i, column= 1).value == None:
            max_row = i
            break
    return max_row-1

### get ticker list from the first row

ticker_list = []

for row in ws.iter_rows(min_row=1, max_row=1, min_col=2, max_col=get_maximum_cols()):
    for cell in row:
        ticker_list.append(cell.value)

### load tickers

loaded_tickers = []

for obj in ticker_list:
    ticker = finvizfinance(obj).ticker_fundament()
    loaded_tickers.append(ticker)

### get data for each ticker

all_ticker_data = []

for i in range(0, len(loaded_tickers)):
    all_ticker_data.extend((
                            loaded_tickers[i].get('Price', 'N/A'),
                            loaded_tickers[i].get('Dividend', 'N/A'),
                            loaded_tickers[i].get('Dividend %', 'N/A'),
                            loaded_tickers[i].get('P/E', 'N/A'),
                            loaded_tickers[i].get('Forward P/E', 'N/A'),
                            loaded_tickers[i].get('Debt/Eq', 'N/A'),
                            loaded_tickers[i].get('SMA20', 'N/A'),
                            loaded_tickers[i].get('EPS (ttm)', 'N/A'),
                            loaded_tickers[i].get('EPS next 5Y', 'N/A'),
                            loaded_tickers[i].get('Earnings', 'N/A'),
                            loaded_tickers[i].get('SMA50', 'N/A'),
                            loaded_tickers[i].get('Payout', 'N/A'),
                            loaded_tickers[i].get('SMA200', 'N/A'),
                            loaded_tickers[i].get('Short Float / Ratio', 'N/A'),
                            loaded_tickers[i].get('Target Price', 'N/A'),
                            str(float(loaded_tickers[i].get('52W Range To', 'N/A')) - float(loaded_tickers[i].get('52W Range From', 'N/A'))),
                            loaded_tickers[i].get('52W High', 'N/A'),
                            loaded_tickers[i].get('52W Low', 'N/A'),
                            loaded_tickers[i].get('RSI (14)', 'N/A'),
                            loaded_tickers[i].get('Rel Volume', 'N/A'),
                            loaded_tickers[i].get('Perf Week', 'N/A'),
                            loaded_tickers[i].get('Perf Month', 'N/A'),
                            loaded_tickers[i].get('Perf Quarter', 'N/A'),
                            loaded_tickers[i].get('Perf Half Y', 'N/A'),
                            loaded_tickers[i].get('Perf Year', 'N/A'),
                            loaded_tickers[i].get('Perf YTD', 'N/A'),
                            loaded_tickers[i].get('Beta', 'N/A')
                            ))

### format tickers

all_tickers_formatted = []
for item in all_ticker_data:
    ftick = item.replace(".",",")
    all_tickers_formatted.append(ftick)

### put data into worksheet

i = 0

for col in ws.iter_cols(min_row=2, max_row=get_maximum_rows(), min_col=2, max_col=get_maximum_cols()):
    for cell in col:
        cell.value = all_tickers_formatted[i]
        i += 1

### save workbook

wb.save('finvizfinance_workbook.xlsx')